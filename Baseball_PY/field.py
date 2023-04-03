import math
from tkinter import *
from openpyxl import *
import sys
import subprocess
import runpy

wb = load_workbook('file.xlsx')     #opens an excel workbook with the name in quotes
ws = wb.active

active = True
WIDTH = 800
HEIGHT = 800
CANVAS_MID_X = WIDTH / 2
CANVAS_MID_Y = HEIGHT / 2
SIDE = WIDTH / 4

root = Tk()
canvas = Canvas(root, bg="black", height=HEIGHT, width=WIDTH)
canvas.pack()

vertices = [
        [CANVAS_MID_X - SIDE / 2, CANVAS_MID_Y - SIDE / 2],
        [CANVAS_MID_X + SIDE / 2, CANVAS_MID_Y - SIDE / 2],
        [CANVAS_MID_X + SIDE / 2, CANVAS_MID_Y + SIDE / 2],
        [CANVAS_MID_X - SIDE / 2, CANVAS_MID_Y + SIDE / 2],
    ]

def rotate(points, angle, center):
    angle = math.radians(angle)
    cos_val = math.cos(angle)
    sin_val = math.sin(angle)
    cx, cy = center
    new_points = []
    for x_old, y_old in points:
            x_old -= cx
            y_old -= cy
            x_new = x_old * cos_val - y_old * sin_val
            y_new = x_old * sin_val + y_old * cos_val
            new_points.append([x_new + cx, y_new + cy])
    return new_points

def draw_square(points, color):
    canvas.create_polygon(points, fill=color)

def test():
    old_vertices = [[150, 150], [250, 150], [250, 250], [150, 250]]
    print("vertices: ", vertices, "should be: ", old_vertices)
    print(vertices == old_vertices)

def getcoord(event):
    global Cx, Cy
    Cx = event.x
    Cy = event.y
    #canvas.create_oval(Cx-4, Cy-4, Cx + 4, Cy + 4)
    print('X = ', Cx, '   Y=  ', Cy)
    ws.cell(column=7, row=ws.max_row, value=Cx)
    ws.cell(column=8, row=ws.max_row, value=Cy)
    wb.save("file.xlsx")
    wb.close()
    root.destroy()
    runpy.run_path('input.py')

#    os.system("Spraychart.xlsx")

def quit(event):
    sys.exit()





center = (CANVAS_MID_X, CANVAS_MID_Y)
vertices = [
        [CANVAS_MID_X - SIDE, CANVAS_MID_Y - SIDE],
        [CANVAS_MID_X + SIDE, CANVAS_MID_Y - SIDE],
        [CANVAS_MID_X + SIDE, CANVAS_MID_Y + SIDE],
        [CANVAS_MID_X - SIDE, CANVAS_MID_Y + SIDE],
    ]
out_square = rotate(vertices, 45, center)
canvas.create_arc(116, 100, 684, 700, start=0, extent=180, fill="green")
draw_square(out_square, 'green')
vertices = [
        [100 + CANVAS_MID_X - SIDE / 2, 100 + CANVAS_MID_Y - SIDE / 2],
        [100 + CANVAS_MID_X + SIDE / 2, 100 + CANVAS_MID_Y - SIDE / 2],
        [100 + CANVAS_MID_X + SIDE / 2, 100 + CANVAS_MID_Y + SIDE / 2],
        [100 + CANVAS_MID_X - SIDE / 2, 100 + CANVAS_MID_Y + SIDE / 2],
    ]
inf_square = rotate(vertices, 45, center)
draw_square(inf_square,'brown')

canvas.bind('<Button-1>', getcoord)
canvas.bind('<Double-1>', quit)

mainloop()