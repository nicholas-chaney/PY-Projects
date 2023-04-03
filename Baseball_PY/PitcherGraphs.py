import tkinter as tk
from tkinter import simpledialog
from matplotlib import pyplot as plt
from openpyxl import *
from collections import Counter
import subprocess
import runpy

wb = load_workbook('file.xlsx')
ws = wb.active

pitchtype = []
pitchspeed = []
rows = ws.max_row


#def PitchByVelo():
#    for i in range(2, rows):
#        datax = ws.cell(row=i+1, column=2).value
#        datay = float(ws.cell(row=i+1, column=4).value)
#        #datay = int(datay)
#        datay = round(datay)
#        pitchtype.append(datax)
#        pitchspeed.append(int(datay))

#    c = Counter(zip(pitchtype, pitchspeed))
#    s = [15*c[(xx, yy)] for xx, yy in zip(pitchtype, pitchspeed)]

#    plt.scatter(pitchtype, pitchspeed, s=s)
#    plt.ylim(72, 98)
#    plt.yticks([72, 74, 76, 78, 80, 82, 84, 86, 88, 90, 92, 94, 96, 98])
#    plt.grid()
#    plt.title('Average Pitch Speed')
#    plt.ylabel('Speed (MPH)')
#    plt.xlabel('Pitch Type')
#    plt.show()


def NumOfTimesPitchThrown():
    PitchesThrown = []
    numoftimes = [0,0,0]
    temp = []
    for i in range(2, ws.max_row+1):
        PitchesThrown.append(str(ws.cell(row=i + 1, column=2).value))
    for j in PitchesThrown:
        if j not in temp:
            temp.append(j)
    for j in range(0, len(temp)-1):
        numoftimes[j] = PitchesThrown.count(temp[j])
    label = [value for value in temp if value != "None"]

    plt.pie(numoftimes,labels=label, autopct='%1.1f%%')
    plt.axis('equal')
    plt.title('Percentage of Pitches Thrown')
    plt.show()


def FirstPitchByPitch():
    pitch0 = ['FB', 'SL', 'CH']
    temp = []
    num = [0, 0, 0]
    for l in range(2, ws.max_row):
        if str(ws.cell(row=l+1, column=5).value) == "0-0":
            if str(ws.cell(row=l+1, column=2).value) == 'FB':
                num[0] = num[0] + 1
            elif str(ws.cell(row=l+1, column=2).value) == 'SL':
                num[1] = num[1] + 1
            elif str(ws.cell(row=l+1, column=2).value) == 'CH':
                num[2] = num[2] + 1

    plt.scatter(pitch0, num)
    plt.yticks([0, 2, 4, 6, 8, 10, 12])
    plt.title('Pitches Thrown in 0-0 Count')
    plt.ylabel('# of Times Thrown')
    plt.xlabel('Pitch Type')
    plt.grid()
    plt.show()


def StrikeLocationChart():
    Locations = ["HI", "HM", "HA", "MI", "MM", "MA", "LI", "LM", "LA"]
    LocCount = [0, 0, 0, 0, 0, 0, 0, 0, 0]

    for l in range(2, rows):
        if str(ws.cell(row=l + 1, column=6).value) == "strike" or "foul" or "ground out" or "strike out" or "single" or "double" or "line out" or "double play":
            if str(ws.cell(row=l + 1, column=3).value) == "HI":
                LocCount[0] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "HM":
                LocCount[1] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "HA":
                LocCount[2] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "MI":
                LocCount[3] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "MM":
                LocCount[4] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "MA":
                LocCount[5] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "LI":
                LocCount[6] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "LM":
                LocCount[7] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "LA":
                LocCount[8] += 1

    plt.scatter(Locations, LocCount)
    plt.yticks([0, 2, 4, 6, 8, 10, 12])
    plt.title('Location of strikes thrown')
    plt.ylabel('# of Times Thrown')
    plt.xlabel('location')
    plt.grid()
    plt.show()


NumOfTimesPitchThrown()
StrikeLocationChart()
#PitchByVelo()
FirstPitchByPitch()
