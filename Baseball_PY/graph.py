import tkinter as tk
from tkinter import simpledialog
from matplotlib import pyplot as plt
from openpyxl import *
from collections import Counter
import subprocess
import runpy

wb = load_workbook('file.xlsx')
ws = wb.active

pitchtype = []
pitchspeed = []
rows = ws.max_row

def findbatterresults():
    #global tempxcords, tempycords
    root = tk.Tk()
    root.withdraw()
    batterinput = simpledialog.askstring(title="Find Batter History", prompt="What batter do you want to see the history of?")
    pitchthrown = []
    location = []
    outcome = []
    count = []
    tempxcords = []
    tempycords = []

    for j in range(3, ws.max_row+1):

        if str(ws.cell(row=j, column=1).value) == batterinput:
            pitchthrown.append(str(ws.cell(row=j, column=2).value))
            location.append(str(ws.cell(row=j, column=3).value))
            outcome.append(str(ws.cell(row=j, column=6).value))
            count.append(str(ws.cell(row=j, column=5).value))
            tempxcords.append(str(ws.cell(row=j, column=7).value))
            tempycords.append(str(ws.cell(row=j, column=8).value))
    print(tempxcords)
    print(tempycords)
    tempxcords_updated = [value for value in tempxcords if value != "None"]
    tempycords_updated = [value for value in tempycords if value != "None"]
    print(tempxcords_updated)
    for x in 1,len(tempxcords_updated):
        ws.cell(row=x, column=9, value=tempxcords_updated[x-1])
        ws.cell(row=x, column=10, value=tempycords_updated[x-1])
    wb.save('file.xlsx')
    runpy.run_path(path_name="SingleBatterSprayChart.py")

def PitchByVelo():
    for i in range(2, rows):
        datax = ws.cell(row=i+1, column=2).value
        datay = float(ws.cell(row=i+1, column=4).value)
        #datay = int(datay)
        datay = round(datay)
        pitchtype.append(datax)
        pitchspeed.append(int(datay))

    c = Counter(zip(pitchtype, pitchspeed))
    s = [15*c[(xx, yy)] for xx, yy in zip(pitchtype, pitchspeed)]

    plt.scatter(pitchtype, pitchspeed, s=s)
    plt.ylim(72, 98)
    plt.yticks([72, 74, 76, 78, 80, 82, 84, 86, 88, 90, 92, 94, 96, 98])
    plt.grid()
    plt.title('Average Pitch Speed')
    plt.ylabel('Speed (MPH)')
    plt.xlabel('Pitch Type')
    plt.show()


def NumOfTimesPitchThrown():
    PitchesThrown = []
    numoftimes = [0,0,0,0,0]
    temp = []
    for i in range(2, rows):
        PitchesThrown.append(str(ws.cell(row=i + 1, column=2).value))
    for j in PitchesThrown:
        if j not in temp:
            temp.append(j)
    for j in range(0, len(temp)):
        numoftimes[j] = PitchesThrown.count(temp[j])

    plt.pie(numoftimes, labels=temp, autopct='%1.1f%%')
    plt.axis('equal')
    plt.title('Percentage of Pitches Thrown')
    plt.show()


def FirstPitchByPitch():
    pitch0 = []
    temp = []
    num = [0, 0, 0, 0]
    for l in range(2, rows):
        if str(ws.cell(row=l+1, column=5).value) == "0-0":
            pitch0.append(str(ws.cell(row=l+1, column=2).value))
    for j in pitch0:
        if j not in temp:
            temp.append(j)
    for j in range(0, len(temp)):
        num[j] = pitch0.count(temp[j])

    plt.scatter(temp, num)
    plt.yticks([0, 2, 4, 6, 8, 10, 12])
    plt.title('Pitches Thrown in 0-0 Count')
    plt.ylabel('# of Times Thrown')
    plt.xlabel('Pitch Type')
    plt.grid()
    plt.show()


def StrikeLocationChart():
    Locations = ["HI", "HM", "HA", "MI", "MM", "MA", "LI", "LM", "LA"]
    LocCount = [0, 0, 0, 0, 0, 0, 0, 0, 0]

    for l in range(2, rows):
        if str(ws.cell(row=l + 1, column=6).value) == "strike" or "foul" or "ground out" or "strike out" or "single" or "double" or "line out" or "double play":
            if str(ws.cell(row=l + 1, column=3).value) == "HI":
                LocCount[0] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "HM":
                LocCount[1] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "HA":
                LocCount[2] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "MI":
                LocCount[3] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "MM":
                LocCount[4] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "MA":
                LocCount[5] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "LI":
                LocCount[6] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "LM":
                LocCount[7] += 1
            if str(ws.cell(row=l + 1, column=3).value) == "LA":
                LocCount[8] += 1

    plt.scatter(Locations, LocCount)
    plt.yticks([0, 2, 4, 6, 8, 10, 12])
    plt.title('Location of strikes thrown')
    plt.ylabel('# of Times Thrown')
    plt.xlabel('location')
    plt.grid()
    plt.show()

findbatterresults()
#NumOfTimesPitchThrown()
#StrikeLocationChart()
#PitchByVelo()
#FirstPitchByPitch()
