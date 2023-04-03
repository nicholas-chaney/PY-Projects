import math
import sys
from tkinter import *
from openpyxl import *
import os

wb = load_workbook('file.xlsx')     #opens an excel workbook with the name in quotes
ws = wb.active
rows = ws.max_row

active = True
WIDTH = 800
HEIGHT = 800
CANVAS_MID_X = WIDTH / 2
CANVAS_MID_Y = HEIGHT / 2
SIDE = WIDTH / 4

root = Tk()
canvas = Canvas(root, bg="black", height=HEIGHT, width=WIDTH)
canvas.pack()

vertices = [
        [CANVAS_MID_X - SIDE / 2, CANVAS_MID_Y - SIDE / 2],
        [CANVAS_MID_X + SIDE / 2, CANVAS_MID_Y - SIDE / 2],
        [CANVAS_MID_X + SIDE / 2, CANVAS_MID_Y + SIDE / 2],
        [CANVAS_MID_X - SIDE / 2, CANVAS_MID_Y + SIDE / 2],
    ]

def rotate(points, angle, center):
    angle = math.radians(angle)
    cos_val = math.cos(angle)
    sin_val = math.sin(angle)
    cx, cy = center
    new_points = []
    for x_old, y_old in points:
            x_old -= cx
            y_old -= cy
            x_new = x_old * cos_val - y_old * sin_val
            y_new = x_old * sin_val + y_old * cos_val
            new_points.append([x_new + cx, y_new + cy])
    return new_points

def draw_square(points, color):
    canvas.create_polygon(points, fill=color)

def test():
    old_vertices = [[150, 150], [250, 150], [250, 250], [150, 250]]
    print("vertices: ", vertices, "should be: ", old_vertices)
    print(vertices == old_vertices)



def quit(event):
    sys.exit()



center = (CANVAS_MID_X, CANVAS_MID_Y)
vertices = [
        [CANVAS_MID_X - SIDE, CANVAS_MID_Y - SIDE],
        [CANVAS_MID_X + SIDE, CANVAS_MID_Y - SIDE],
        [CANVAS_MID_X + SIDE, CANVAS_MID_Y + SIDE],
        [CANVAS_MID_X - SIDE, CANVAS_MID_Y + SIDE],
    ]
out_square = rotate(vertices, 45, center)
canvas.create_arc(116, 100, 684, 700, start=0, extent=180, fill="green")
draw_square(out_square, 'green')
vertices = [
        [100 + CANVAS_MID_X - SIDE / 2, 100 + CANVAS_MID_Y - SIDE / 2],
        [100 + CANVAS_MID_X + SIDE / 2, 100 + CANVAS_MID_Y - SIDE / 2],
        [100 + CANVAS_MID_X + SIDE / 2, 100 + CANVAS_MID_Y + SIDE / 2],
        [100 + CANVAS_MID_X - SIDE / 2, 100 + CANVAS_MID_Y + SIDE / 2],
    ]
inf_square = rotate(vertices, 45, center)
draw_square(inf_square,'brown')

for i in range(3,ws.max_row + 1):
    xcords = []
    ycords = []

    xcord = ws.cell(row=i, column=7).value
    ycord = ws.cell(row=i, column=8).value
    if xcord == None:
        print("There are no Hits!!!")
    else:
        print(type(xcord))
        canvas.create_oval(xcord - 4, ycord - 4, xcord + 4, ycord + 4, fill='red')

mainloop()